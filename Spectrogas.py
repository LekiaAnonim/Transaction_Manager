#Creating the Graphical User Interface
#==========================================================================================================================
from tkinter import*

import pandas as pd
from tkinter import messagebox
from datetime import datetime
#import dash
#import dash_core_components as dcc
#import dash_html_components as html
import pandas as pd
import numpy as np
#from dash.dependencies import Output, Input
import os
import webbrowser
#from PIL import Image, ImageTk
root = Tk()
root.iconbitmap('TM.ico')
root.title('Transaction Manager')
root.resizable(0,0)


loginframe=LabelFrame(root, text='Login Page', bg='#afd6f5')
loginframe.place(relwidth=0.88,relheight=0.7,rely=0.1, relx=0.05)
password_label = Label(loginframe, text= 'Password', bg='#afd6f5').place(rely=0.3, relx = 0.05)

password_entry = Entry(loginframe, show='*')
password_entry.place(rely=0.3, relx = 0.3)

Password = 'pilotx_lab'

def login():
    password = password_entry.get()
    if str(password) == Password:
        global app
        def app():
            
            def visual():
                from subprocess import Popen
            
                Popen('python spectro.py')
                url = "http://localhost:8050/"
                webbrowser.open_new(url)
                
            app = Tk()
            app.iconbitmap('TM.ico')
            app.title('Transaction Manager')
            app.resizable(0,0)
            
            def on_enter(e):
                e.widget['background'] = '#03a9f4'

            def on_leave(e):
                e.widget['background'] = 'SystemButtonFace'
                        
                
            from tkinter import ttk
            tabControl = ttk.Notebook(app)
            
            tab1 = ttk.Frame(tabControl)
            tab2 = ttk.Frame(tabControl)
            tab3 = ttk.Frame(tabControl)
            
            tabControl.add(tab1, text ='Transaction Entry')
            tabControl.add(tab2, text ='Transacion Analysis')
            tabControl.add(tab3, text ='Transacion Visualization')
            tabControl.place(relheight=1, relwidth=1,rely=0, relx=0)
            
            # Creating frames
            NumberOfKg_frame = LabelFrame(tab1, text = 'Number of Kg Purchased')
            NumberOfKg_frame.place(relheight=0.4, relwidth=0.34,rely=0.23, relx=0.28)
            Price_frame = LabelFrame(tab1, text = 'Purchased Price')
            Price_frame.place(relheight=0.4, relwidth=0.34,rely=0.23, relx=0.64)
            Status_frame = LabelFrame(tab1, text = 'Status')
            Status_frame.place(relheight=0.2, relwidth=0.99,rely=0.65, relx=0.01)
            Analysis_frame = LabelFrame(tab2, text='Analysis Result')
            Analysis_frame.place(relheight=0.23,relwidth=0.6,rely=0.19,relx=0.01)
            Search_frame = LabelFrame(tab2, text='Search Result')
            Search_frame.place(relheight=0.4,relwidth=1,rely=0.52,relx=0.01)

            
            
            def NumberOfKg():
                Price_Enter=Price_EnterEntry.get()
                PricePerKg=PricePerKgEntry.get()
                Kg = "%8.2f" %  (float(Price_Enter)/float(PricePerKg))
                NumberOfKg_CalculateLabel = Label(NumberOfKg_frame,text=Kg, bg='white')
                NumberOfKg_CalculateLabel.place(relheight=0.25,relwidth=0.5,rely=0.6,relx=0.45)
                return NumberOfKg_CalculateLabel
            NumberOfKg_CalculateButton = Button(NumberOfKg_frame,text='Calculate',relief=FLAT,cursor='hand2',command=NumberOfKg)
            NumberOfKg_CalculateButton.place(rely=0.6,relx = 0.01)
            NumberOfKg_CalculateButton.bind("<Enter>", on_enter)
            NumberOfKg_CalculateButton.bind("<Leave>", on_leave)

            def PurchasePrice():
                PricePerKg=PricePerKgEntry.get()
                NumberOfKg_Enter=NumberOfKg_EnterEntry.get()
                Price = "%8.2f" % (float(PricePerKg)*float(NumberOfKg_Enter))
                Price_CalculateLabel = Label(Price_frame,text=  Price, bg='white')
                Price_CalculateLabel.place(relheight=0.25,relwidth=0.5,rely=0.6,relx=0.45)
                return Price_CalculateLabel
            Price_CalculateButton = Button(Price_frame,text='Calculate',relief=FLAT,cursor='hand2',command=PurchasePrice)
            Price_CalculateButton.place(rely=0.6,relx = 0.01)
            Price_CalculateButton.bind("<Enter>", on_enter)
            Price_CalculateButton.bind("<Leave>", on_leave)
            


            def Balance():
                AmountPaid = AmountPaidEntry.get()
                Price_Enter = Price_EnterEntry.get()
                balances = "%8.2f" % (float(AmountPaid)-float(Price_Enter))
                BalanceLabel = Label(Status_frame, text = balances, bg='white')
                BalanceLabel.place(relheight=0.3,relwidth=0.15,rely=0.3,relx=0.13)
                return BalanceLabel
            
                        
            BalanceButton = Button(Status_frame,text='Balance',relief=FLAT,cursor='hand2',command=Balance)
            BalanceButton.place(rely=0.3,relx=0.01)
            BalanceButton.bind("<Enter>", on_enter)
            BalanceButton.bind("<Leave>", on_leave)

            now = datetime.now()
            # dd/mm/YY H:M:S
            date_time = now.strftime("%B %d, %Y")
            date_and_time=now.strftime("%B %d, %Y | %H:%M:%S")
            
            from openpyxl import load_workbook
            def append_df_to_excel(filename, sheet_name='Sheet1', startrow=None,
                                   truncate_sheet=False, 
                                   **to_excel_kwargs):
                
                
                Price_Enter=Price_EnterEntry.get()
                PricePerKg=PricePerKgEntry.get()
                NumberOfKg_Enter=NumberOfKg_EnterEntry.get()
                Name = NameEntry.get()
                if len(PhoneEntry.get())==11:
                    Phone = PhoneEntry.get()
                else:
                    messagebox.showwarning('Warning',"Check length of Number")
                Debt = DebtEntry.get()
                Credit = CreditEntry.get()
                Clear_Option = ClearOption()
                data = [[date_time,Name,str(Phone),str(Price_Enter),str(NumberOfKg_Enter),Clear_Option,str(Debt),str(Credit)]]
                DataTable = pd.DataFrame(data, columns=['Date','Name of Customer','Phone Number','Amount Paid','Number of Kg','Status','Debt','Credit'])
                messagebox.showinfo('',"Successful!")
                
                try:
                    from openpyxl import load_workbook
            
                    # ignore [engine] parameter if it was passed
                    if 'engine' in to_excel_kwargs:
                        to_excel_kwargs.pop('engine')
            
                    writer = pd.ExcelWriter(filename, engine='openpyxl')
            
            
                    try:
                        # try to open an existing workbook
                        writer.book = load_workbook(filename)
            
                        # get the last row in the existing Excel sheet
                        # if it was not specified explicitly
                        if startrow is None and sheet_name in writer.book.sheetnames:
                            startrow = writer.book[sheet_name].max_row
            
                        # truncate sheet
                        if truncate_sheet and sheet_name in writer.book.sheetnames:
                            # index of [sheet_name] sheet
                            idx = writer.book.sheetnames.index(sheet_name)
                            # remove [sheet_name]
                            writer.book.remove(writer.book.worksheets[idx])
                            # create an empty sheet [sheet_name] using old index
                            writer.book.create_sheet(sheet_name, idx)
            
                        # copy existing sheets
                        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
                    except FileNotFoundError:
                        # file does not exist yet, we will create it
                        pass
            
                    if startrow is None:
                        startrow = 0
            
                    # write out the new sheet
                    DataTable.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)
            
                    # save the workbook
                    writer.save()
                except PermissionError:
                    messagebox.showerror("Information", "Workbook is open. Please close workbook!")
                    return None
            
            
            def AnalysisBydate():
                try:
                    date = DateEntry.get()
                    analysis_data = pd.read_excel('DataBase.xlsx',parse_dates=True,index_col='Date')
                    selected_data = analysis_data.loc[str(date)]
                    Revenue = selected_data['Amount Paid'].sum()
                    RevenueResult = Label(Analysis_frame,text=str(Revenue), bg='white')
                    RevenueResult.place(relwidth=0.27,rely=0.01, relx=0.2)
                    kg_sold = selected_data['Number of Kg'].sum()
                    Kg_SoldResult = Label(Analysis_frame,text=str(kg_sold), bg='white')
                    Kg_SoldResult.place(relwidth=0.2,rely=0.01, relx=0.72)
                    cleared =  len(selected_data[selected_data.Status=='Cleared'])
                    ClearResult = Label(Analysis_frame,text=str(cleared), bg='white')
                    ClearResult.place(relwidth=0.2,rely=0.4, relx=0.19)
                    notcleared = len(selected_data[selected_data.Status =='Not Cleared'])
                    NotClearResult = Label(Analysis_frame,text= str(notcleared), bg='white')
                    NotClearResult.place(relwidth=0.2,rely=0.4, relx=0.72)
                    return( RevenueResult,Kg_SoldResult,ClearResult,NotClearResult)
                except KeyError:
                    messagebox.showerror("Error", "Date not available!")
                    return None
            
            def Searchresult():
                global Searchtable
                searchby = search_n.get()
                search = SearchEntry.get()
                generaldata = pd.read_excel('DataBase.xlsx').astype(str)
                searched_data = generaldata[generaldata[str(searchby)]==search]
            
                Searchtable["column"]=list(searched_data.columns)
                Searchtable['show']='headings'
                for column in Searchtable['columns']:
                    Searchtable.heading(column,text=column)
                searched_data_rows = searched_data.to_numpy().tolist()
                for row in searched_data_rows:
                    
                    Searchtable.insert("","end", values=row)
                    #return None
            
            def Clearsearch():
                Searchtable.delete(*Searchtable.get_children())
            ClearSearchButton = Button(tab2, text='Clear Search',relief=FLAT,fg='white',bg='red', command= Clearsearch)
            ClearSearchButton.place(rely = 0.92, relx=0.1)
            
            NumberOfKg_EnterLabel  = Label(NumberOfKg_frame,text = 'Enter:').place(rely=0.1, relx = 0.01)
            NumberOfKg_EnterEntry = Entry(NumberOfKg_frame)
            NumberOfKg_EnterEntry.place(relwidth = 0.65, rely=0.1,relx = 0.3)
                        
            Price_EnterLabel  = Label(Price_frame,text = 'Enter:').place(rely=0.1, relx = 0.01)

                        
            def ClearOption(): 
                return v.get()
            
            v = StringVar()
            StatusRadioButton1 = Radiobutton(Status_frame,text='Cleared',variable=v,value='Cleared').place(rely=0.08,relx=0.4)
            StatusRadioButton2 = Radiobutton(Status_frame,text='Not Cleared',variable=v,value='Not Cleared').place(rely=0.4,relx=0.4)
            DebtLabel = Label(Status_frame, text='Debt').place(rely=0.08, relx=0.65)
            CreditLabel = Label(Status_frame, text='credit').place(rely=0.45, relx=0.65)
            DebtEntry = Entry(Status_frame)
            DebtEntry.place(relwidth=0.2,rely=0.08, relx = 0.75)
            CreditEntry = Entry(Status_frame)
            CreditEntry.place(relwidth=0.2,rely=0.45, relx = 0.75)
            
            PricePerKgLabel = Label(tab1, text = 'Price per kg:')
            PricePerKgLabel.place(rely=0.23, relx=0.01)
            PricePerKgEntry = Entry(tab1)
            PricePerKgEntry.place(rely=0.3, relx=0.01)
            
            AmountPaidLabel = Label(tab1, text = 'Amount Paid:')
            AmountPaidLabel.place(rely=0.41, relx=0.01)
            Price_EnterEntry = Entry(Price_frame)
            Price_EnterEntry.place(relwidth = 0.65, rely=0.1,relx = 0.3)
            
            AmountPaidEntry = Entry(tab1)
            AmountPaidEntry.place(rely=0.48, relx=0.01)

            NameLabel = Label(tab1, text = 'Name of customer')
            NameLabel.place(rely=0.03,relx=0.01)
            NameEntry = Entry(tab1)
            NameEntry.place(relwidth=0.4,rely=0.1,relx=0.01)
            
            PhoneLabel = Label(tab1, text = 'Phone number')
            PhoneLabel.place(rely=0.03,relx=0.6)
            PhoneEntry = Entry(tab1)
            PhoneEntry.place(rely=0.1,relx=0.6)
            ApplyButton = Button(tab1, text = 'Apply',fg='white',relief=FLAT, bg = '#03a9f4',cursor='hand2', command= lambda: append_df_to_excel('DataBase.xlsx',header=None, sheet_name='Sheet1', index=False, startrow=None))
            ApplyButton.place(relwidth =0.2,rely=0.9, relx =0.1)
            
            Tab2_Info = Label(tab2,text='Instruction : Follow this date format to analyse your data. Example, Year Format : 2021')
            Tab2_Info.place(rely=0.001,relx = 0.01 )
            
            Day_format= Label(tab2, text = 'Day Format : August 10, 2021 or 2021-08-10, and Month Format : August 2021 or 2021-08')
            Day_format.place(rely=0.05,relx = 0.01 )
            
            DateLabel = Label(tab2, text = 'Enter Date:')
            DateLabel.place(rely=0.12, relx=0.01)
            DateEntry = Entry(tab2)
            DateEntry.place(relheight=0.07,rely=0.12,relx=0.17)
            Select_date = DateEntry
            
            RevenueLabel = Label(Analysis_frame,text='Revenue :')
            RevenueLabel.place(rely=0.01, relx=0.01)
            
            Kg_SoldLabel = Label(Analysis_frame,text='Kg Sold :')
            Kg_SoldLabel.place(rely=0.01, relx=0.55)
            ClearLabel = Label(Analysis_frame,text='Cleared :')
            ClearLabel.place(rely=0.4, relx=0.01)
            NotClearLabel = Label(Analysis_frame,text='Not Cleared :')
            NotClearLabel.place(rely=0.4, relx=0.47)
            
            SearchInfo_label1 = Label(tab2,text='Serach By options include :')
            SearchInfo_label1.place(rely=0.19,relx=0.61)
            SearchInfo_label2 = Label(tab2,text='- Name of Customer')
            SearchInfo_label2.place(rely=0.24,relx=0.62)
            SearchInfo_label3 = Label(tab2,text='- Date (e.g. February 17, 2021)')
            SearchInfo_label3.place(rely=0.29,relx=0.62)
            SearchInfo_label2 = Label(tab2,text='- Phone Number (e.g. 7032033155)')
            SearchInfo_label2.place(rely=0.34,relx=0.62)
            
            SearchbyLabel = Label(tab2, text = ' Search By :')
            SearchbyLabel.place(rely=0.46, relx=0.01)
            # Combobox creation 
            search_n = StringVar() 
            Searchbychoice = ttk.Combobox(tab2, width = 27, textvariable = search_n) 
            
            # Adding combobox drop down list 
            Searchbychoice['values'] = ('Name of Customer',  
                                      'Date', 
                                      'Phone Number',
                                      'Status') 
            
            Searchbychoice.place(relheight=0.07,relwidth = 0.3,rely=0.45, relx=0.15) 
            Searchbychoice.current() 
            
            SearchLabel = Label(tab2, text='Search : ')
            SearchLabel.place(rely=0.46, relx=0.49)
            SearchEntry = Entry(tab2)
            SearchEntry.place(relheight=0.07,relwidth = 0.3,rely=0.45, relx=0.59)
            
            Searchtable = ttk.Treeview(Search_frame)
            Searchtable.place(relheight=1, relwidth=1,rely=0, relx=0)
            
            treescrolly = Scrollbar(Search_frame, orient='vertical',command=Searchtable.yview)
            treescrollx = Scrollbar(Search_frame, orient='horizontal',command=Searchtable.xview)
            Searchtable.configure(xscrollcommand=treescrollx.set)
            Searchtable.configure(yscrollcommand=treescrolly.set)
            treescrollx.pack(side='bottom', fill='x')
            treescrolly.pack(side='right', fill='y')
            
            visualLabel = Label(tab3, text='Click on the button to visualize your transaction!')
            visualLabel.place(rely=0.02,relx=0.01)

            
            Ok_button = Button(tab2,text='OK',relief=FLAT, bg='#8bc34a',cursor='hand2', command=AnalysisBydate)
            Ok_button.place(rely=0.12,relx=0.44)
            
            
            SearchButton = Button(tab2, text='OK', bg='#8bc34a',relief=FLAT,cursor='hand2',command=Searchresult)
            SearchButton.place(relheight=0.07,relwidth = 0.08,rely=0.45, relx=0.9)
            
                        #photo = PhotoImage(file = "visual.png")
            visualizationbutton = Button(tab3, text='Visualize',relief=FLAT,cursor='hand2', command= visual)
            visualizationbutton.place(relheight=0.2, relwidth=0.2, rely=0.1, relx=0.1)
            visualizationbutton.bind("<Enter>", on_enter)
            visualizationbutton.bind("<Leave>", on_leave)
            
            Copywrite = Label(app, text='Developer : PilotX Lab', fg='white', bg='black').place(rely=0.95, relx=0.75)
            
            def print_slip():
                f = open('Print_slip.txt', 'w')
                f.write('DATE AND TIME: '+ str(date_and_time)+'\n')
                f.write('\n')
                f.write('TRANSACTION DETAILS \n')
                f.write('Name of Customer: '+ str(NameEntry.get()) + '\n')
                f.write('Phone Number: '+ str(PhoneEntry.get())+'\n')
                price_enter = Price_EnterEntry.get()
                purchase_price = "%8.2f" % float(price_enter)
                f.write('Price Per Kg: '+ str(purchase_price)+' Naira' +'\n')
                amountpaid = AmountPaidEntry.get()
                moneypaid = "%8.2f" % float(amountpaid)
                f.write('Amount Paid: '+ str(moneypaid)+ ' Naira'+'\n')
                f.write('Number of Kg Purchased: '+ str(NumberOfKg_EnterEntry.get()) + 'kg'+'\n')
                f.write('Purchased Price: '+ str(Price_EnterEntry.get())+' Naira'+'\n')
                
                
                balance = "%8.2f" % (float(amountpaid)-float(price_enter))
                f.write('Balance: '+ str(balance)+' Naira'+'\n')
                f.write('Debt: '+ str(DebtEntry.get())+'\n')
                f.write('Credit: '+ str(CreditEntry.get())+'\n')
                clear_option = ClearOption()
                f.write('Status: '+ clear_option +'\n')
                f.close()
                os.startfile("Print_slip.txt", "print")
            print_button = Button()
            print_button = Button(tab1, text = 'Print Slip',fg='white',relief=FLAT, bg = '#03a9f4',cursor='hand2', command= print_slip)
            print_button.place(relwidth =0.2,rely=0.9, relx =0.5)

            app.geometry("500x400")
            app.mainloop()

        app=app()
    else:
        app=messagebox.showerror("Information","Wrong Password")
    return app
    
login_button = Button(loginframe, text='Login',relief=FLAT,bg='#ff5722',cursor='hand2',command=login).place(rely=0.5, relx=0.3)        
    
root.geometry("300x200")
root.mainloop()

